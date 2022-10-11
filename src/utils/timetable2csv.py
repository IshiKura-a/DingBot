import re

import openpyxl
import argparse

parser = argparse.ArgumentParser()
parser.add_argument('--input', type=str)
parser.add_argument('--output', type=str, default='timetable.csv')
args = parser.parse_args()


def main():
    book = openpyxl.load_workbook(args.input)
    classes = []
    for sheet in [book[name] for name in book.sheetnames]:
        title = re.compile(r'（(.+)）').findall(sheet.title)[0]
        max_row = sheet.max_row
        max_col = sheet.max_column
        classroom = [
            str(sheet.cell(1, i).value) if sheet.cell(1, i).value is not None else str(sheet.cell(1, i - 1).value)
            for i in range(2, max_col + 1)]
        for i in range(3, max_row + 1, 2):
            time = str(sheet.cell(i, 1).value)
            for j in range(2, max_col + 1):
                if sheet.cell(i, j).value is not None:
                    course = str(sheet.cell(i, j).value)
                    students = str(sheet.cell(i + 1, j).value)
                    classes.append([
                        title, str(sheet.cell(2, j).value), time, classroom[j - 2], course, students
                    ])

    with open('../../input/timetable.csv', 'w', encoding='UTF-8') as f:
        for line in classes:
            print(",".join(line), file=f)
    return


if __name__ == '__main__':
    main()
